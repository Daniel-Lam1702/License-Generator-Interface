#Elaborado por Daniel Eduardo Lam He y Miguel Aguilar
#Fecha de creación: 06/06/2021 2:56 pm
#Fecha de modificación: 11/06/2021 6:24 pm
#Versión: 3.9.5
#Llamada de librerías
from clases import *
from memoriaSecundaria import *
import random
import names
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment,Border,Side
from fpdf import FPDF
from bs4 import BeautifulSoup
import requests
import xml.etree.cElementTree as ET
#Definición de funciones
#Generar XML
def webScrapping():
    """
    Funcionamiento:Extraccion de la informacion de la pagina web
    Entradas:NA
    Salidas:NA
    """
    try:
        html = requests.get("https://practicatest.cr/blog/licencias/tipos-licencia-conducir-costa-rica").text
        soup = BeautifulSoup(html,"lxml")
        licencias = soup.find("div", class_="col-md-12 content-blog")
        listaTipos=[]
        listaSubtipos=[]
        todosReq=[]
        listaComentarios=[]
        def eliminarEspacio(pTexto):
            if "\n" in pTexto:
                return pTexto.replace("\n","")
            else:
                return pTexto
        for tipo in licencias.find_all("h2"):
            listaTipos.append(tipo.text)
        for subtipo in licencias.find_all("h3"):
            listaSubtipos.append(eliminarEspacio(subtipo.text))
        for requisito in licencias.find_all("ul"):
            ListaReq=[]
            reqLi= requisito.find_all("li")
            for req in reqLi:
                if "" != req.text:
                    ListaReq.append(eliminarEspacio(req.text))
            todosReq.append(ListaReq)
        for comentario in licencias.find_all("p"):
            if "" != comentario.text and " " != comentario.text:
                listaComentarios.append(eliminarEspacio(comentario.text))
        print("-------------------------------Tipos---------------------------------------------")
        print(listaTipos)
        print("-------------------------------Subtipos---------------------------------------------")
        print(listaSubtipos)
        print("-------------------------------Requisitos---------------------------------------------")
        print(todosReq)
        print("-------------------------------Comentarios---------------------------------------------")
        print(listaComentarios[4:][:12][:3]+[listaComentarios[4:][:12][3:7]]+listaComentarios[4:][:12][7:12]+listaComentarios[4:][16:])
        generarXML(listaTipos,listaSubtipos,todosReq,listaComentarios[4:][:12][:3]+[listaComentarios[4:][:12][3:7]]+listaComentarios[4:][:12][7:12]+listaComentarios[4:][16:])
        return True
    except:
        return False
def aplicarRequisitos(pArbol,pReq):
    """
    Funcionamiento:Aplicar los requisitos para el xml
    Entradas:
    pArbol
    pReq Requsitos
    Salidas:NA
    """
    for requisito in pReq:
        ET.SubElement(pArbol,"Requisitos").text=requisito
    return
def generarXML(pTipos,pSubtipos,pRequisitos,pComentarios):
    """
    Funcionamiento:Generar el xml
    Entradas:pTipos(str)tipo de licencia
    pSubtipos(str)subtipo de licencia
    pRequisitos(str)los requisitos para cada licencia
    pComentarios(str)El comentario de cada licencia
    Salidas:NA
    """
    licencias = ET.Element("licencias")
    #Añadir del tipo A
    tipoA=ET.SubElement(licencias,"Tipo",licencia=pTipos[0])
    for i in range(3):
        #Añadir subtipo A1,A2,A3
        subtipoA=ET.SubElement(tipoA,"Subtipo",licencia=pSubtipos[i])
        aplicarRequisitos(subtipoA,pRequisitos[i])
        ET.SubElement(subtipoA,"Comentarios").text=pComentarios[i]
    #Añadir del tipo B
    tipoB=ET.SubElement(licencias,"Tipo",licencia=pTipos[1])
    tipoB1=ET.SubElement(tipoB,"Suntipo",licencia=pSubtipos[3])
    aplicarRequisitos(tipoB1,pRequisitos[3])
    for comentario in pComentarios[3]:
        ET.SubElement(tipoB1,"Requisitos").text=comentario
    for i in range(4,7):
        #Añadir subtipo B2,B3,B4
        subtipoB=ET.SubElement(tipoB,"Subtipo",licencia=pSubtipos[i])
        aplicarRequisitos(subtipoB,pRequisitos[i])
        ET.SubElement(subtipoB,"Comentarios").text=pComentarios[i]
    #Añadir del tipo C
    tipoC=ET.SubElement(licencias,"Tipo",licencia=pTipos[2])
    for i in range(7,9):
        subtipoC=ET.SubElement(tipoC,"Subtipo",licencia=pSubtipos[i])
        aplicarRequisitos(subtipoC,pRequisitos[i])
        ET.SubElement(subtipoC,"Comentarios").text=pComentarios[i]
    #Añadir del tipo D
    tipoD=ET.SubElement(licencias,"Tipo",licencia=pTipos[3])
    for i in range(9,12):
        subtipoD=ET.SubElement(tipoD,"Subtipo",licencia=pSubtipos[i])
        aplicarRequisitos(subtipoD,pRequisitos[9])
        ET.SubElement(subtipoD,"Comentarios").text=pComentarios[i]
    #Añadir del tipo E
    j=10
    tipoE=ET.SubElement(licencias,"Tipo",licencia=pTipos[4])
    for i in range(12,14):
        subtipoE=ET.SubElement(tipoE,"Subtipo",licencia=pSubtipos[i])
        aplicarRequisitos(subtipoE,pRequisitos[j])
        j+=1
        ET.SubElement(subtipoE,"Comentarios").text=pComentarios[i]
    archivo=ET.ElementTree(licencias)
    archivo.write("licencias.xml",encoding="utf-8")
    return
#Generar conductores
def verSiMayor25(fechaNacimiento):
    """
    Funcionamiento:Validar la fecha de nacimiento(edad)
    Entradas:fechaNacimiento(str)fecha de nacimiento de la persona
    Salidas:NA
    """
    if datetime.datetime.now().year-int(fechaNacimiento[6:])>25:
        return True
    else:
        return False
def generarCedula():
    """
    Funcionamiento:Generar una cedula aleatoria
    Entradas:NA
    Salidas:NA
    """
    conductores = leer("conductores")
    numero = str(random.randint(100000000,999999999))
    for licencia in conductores:
        if licencia.getCedula() == numero:
            return generarCedula()
    return numero
def generarNombre():
    """
    Funcionamiento:Generar un nombre aleatorio
    Entradas:NA
    Salidas:NA
    """
    return names.get_full_name() + " " + names.get_last_name()
def generarFechaNacimiento():
    """
    Funcionamiento:Generar una fecha de nacimiento
    Entradas:NA
    Salidas:NA
    """
    start_date = datetime.date(datetime.date.today().year-123, 1, 1)
    end_date = datetime.date(datetime.date.today().year-18, datetime.date.today().month, datetime.date.today().day-1)
    time_between_dates = end_date - start_date
    days_between_dates = time_between_dates.days
    random_number_of_days = random.randrange(days_between_dates)
    random_date = start_date + datetime.timedelta(days=random_number_of_days)
    return str(random_date)[8:]+"-"+str(random_date)[5:7]+"-"+str(random_date)[0:4]
def generarFechaExpedicion():
    """
    Funcionamiento:Generar la fecha de expedicion
    Entradas:NA
    Salidas:NA
    """
    fecha=str(datetime.datetime.now())
    return fecha[8:10]+fecha[4:7]+"-"+fecha[0:4]
def generarFechaVencimiento(pFechaNac,pFechaExpedicion):
    """
    Funcionamiento:generar la fecha de vencimiento de la licencia
    Entradas:pFechaNac(str)fecha de nacimiento
    pFechaExpedicion(str)fecha de expedicion
    Salidas:NA
    """
    if verSiMayor25(pFechaNac):
        return pFechaExpedicion[:-4]+str(int(pFechaExpedicion[6:])+5)
    else:
        return pFechaExpedicion[:-4]+str(int(pFechaExpedicion[6:])+3)
def generarTipoLicencia():
    """
    Funcionamiento:Generar los tipos de licencia de forma aleatoria
    Entradas:NA
    Salidas:NA
    """
    #Leer del xml
    try:
        xml = ET.parse("licencias.xml")
        root = xml.getroot()
        listaLicencias=[]
        for tipo in root:
            for subtipo in tipo:
                listaLicencias.append(subtipo.attrib["licencia"][9:11])
        return listaLicencias[random.randint(0,len(listaLicencias)-1)]
    except:
        return "Error"
def generarTipoSangre():
    """
    Funcionamiento:Generar el tipo de sangre de forma aleatoria
    Entradas:NA
    Salidas:NA
    """
    listaSangres=["O+","O-","A+","A-","B+","B-","AB+","AB-"]
    return listaSangres[random.randint(0,len(listaSangres)-1)]
def generarDonador():
    """
    Funcionamiento:Generar de forma aleatoria si la persona es donador
    Entradas:NA
    Salidas:NA
    """
    return bool(random.randint(0,1))
def generarSede(pDigito):
    """
    Funcionamiento:Generar de forma aleatoria la sede
    Entradas:pDigito(int)Digito para el tipo de sede
    Salidas:NA
    """
    sedes={"189":["San Sebastián","Pérez Zeledón"],"2":["Montecillos Alajuela","Tránsito San Ramón","San Carlos"],
    "3":["Tránsito Cartago"],"4":["Barva de Heredia"],"5":["Liberia","Nicoya"],"6":["Chacarita Puntarenas","Río Claro de Golfito"],
    "7":["Guápiles","Terminal de contenedores de SELDECA Limón"]}
    for digito in sedes:
        if pDigito in digito:
            return sedes[digito][(random.randint(0,len(sedes[digito])-1))]
    return
def generarPuntaje():
    """
    Funcionamiento:Generar un puntaje aleatorio
    Entradas:NA
    Salidas:NA
    """
    return random.randint(0,12)
def generarCorreo(pNombre):
    """
    Funcionamiento:Generar un correo de forma aleatoria
    Entradas:pNombre(str)nombre aleatorio
    Salidas:NA
    """
    listaNombre=pNombre.split(" ")
    return listaNombre[1].lower()+listaNombre[2][0].lower()+listaNombre[0][0].lower()+"@gmail.com"
def generarLicencias(pCant):
    """
    Funcionamiento:Genera la licencia con todos los datos requeridos
    Entradas:pCant(int)Cantidad de licencias a crear
    Salidas:NA
    """
    try:
        if generarTipoLicencia()=="Error":
            return "Error"
        lista=leer("conductores")
        for conductor in range(pCant):
            cedula=generarCedula()
            nombre=generarNombre()
            fechaNacimiento=generarFechaNacimiento()
            fechaExpedicion=generarFechaExpedicion()
            fechaVencimiento=generarFechaVencimiento(fechaNacimiento,fechaExpedicion)
            tipoLicencia=generarTipoLicencia()
            tipoSangre=generarTipoSangre()
            donador=generarDonador()
            sede=generarSede(cedula[0])
            puntaje=generarPuntaje()
            correo=generarCorreo(nombre)
            objeto=Conductores(cedula,nombre,fechaNacimiento,fechaExpedicion,fechaVencimiento,tipoLicencia,tipoSangre,donador,sede,puntaje,correo)
            lista.append(objeto)
        grabar("conductores",lista)
        return True
    except:
        return False
#Renovar licencia
def renovarLicencia(pCedula):
    """
    Funcionamiento:Opcion para renovar la licencia
    Entradas:pCedula(str)numero de cedula
    Salidas:NA
    """
    objetos=leer("conductores")
    try:
        for objeto in objetos:
            if objeto.getCedula()==pCedula:
                objeto.setFechaExpedicion(generarFechaExpedicion())
                objeto.setFechaVencimiento(generarFechaVencimiento(objeto.getFechaNacimiento(),generarFechaExpedicion()))
                objeto.setPuntaje(12)
                grabar("conductores",objetos)
                return True
    except:
        return False
#Generar PDF
def decodificarDonadorPdf(booleano):
    """
    Funcionamiento:Dedifica si el conductor es donador o no
    Entradas:booleano(bool) True or False
    Salidas:NA
    """
    if booleano:
        return "Donador"
    else:
        return "No donador"
def generarPdf(cedula):
    """
    Funcionamiento:Generar el pdf con la informacion del conductor
    Entradas:cedula(str)numero de cedula
    Salidas:NA
    """
    try:
        pdf=FPDF('P','mm','Letter')
        pdf.add_page()
        pdf.set_font('times','',16)
        pdf.set_text_color(0,191,255)
        pdf.cell(120,10,"REPUBLICA DE COSTA RICA",ln=True)
        pdf.set_text_color(255, 0, 0)
        pdf.cell(110,10,"Licencia de Conducir",ln=True)
        objetos=leer("conductores")
        for objeto in objetos:
            if objeto.getCedula() == cedula:
                pdf.set_text_color(0, 0, 0)
                pdf.set_font('times','B',16)
                pdf.cell(10,10,"N°:")
                pdf.set_text_color(255, 0, 0)
                pdf.cell(10,10,"CI-"+objeto.getCedula(),ln=True)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(32,10,"Expedición:")
                pdf.set_font('times','',16)
                pdf.cell(32,10,objeto.getFechaExpedicion(),ln=True)
                pdf.set_font('times','B',16)
                pdf.cell(32,10,"Nacimiento: ")
                pdf.set_font('times','',16)
                pdf.cell(32,10,objeto.getFechaNacimiento(),ln=True)
                pdf.set_font('times','B',16)
                pdf.cell(35,10,"Vencimiento: ")
                pdf.set_text_color(255, 0, 0)
                pdf.set_font('times','',16)
                pdf.cell(35,10,objeto.getFechaVencimiento(),ln=True)
                pdf.cell(13,10,"Tipo: ")
                pdf.set_font('times','B',16)
                pdf.cell(13,10,objeto.getTipoLicencia(),ln=True)
                pdf.set_font('times','',16)
                pdf.cell(10,10,decodificarDonadorPdf(objeto.getDonador()),ln=True)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(10,10,"T.S. ")
                pdf.set_text_color(255, 0, 0)
                pdf.cell(10,10,objeto.getTipoSangre(),ln=True)
                pdf.set_text_color(0, 0, 0)
                pdf.set_font('times','B',26)
                pdf.cell(10,10,str(objeto.getNombre()).upper(),ln=True)
                pdf.set_font('times','',16)
                pdf.cell(10,10,str(datetime.datetime.now())[8:10]+str(datetime.datetime.now())[4:7]+"-"+str(datetime.datetime.now())[0:4]+str(datetime.datetime.now())[10:-10]+" "+objeto.getSede(),ln=True)
                pdf.output(objeto.getCedula()+".pdf")
                return True
    except:
        return False
#------------------Reporte de licencias------------
def generarReporteLicencia():
    """
    Funcionamiento: Crear el reporte en excel de las licencias
    Entradas:NA
    Salidas:NA
    """
    try:
        licencia=leer("conductores")
        largoNombre=largoSede="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = "Licencias Activas"
        excelSheet["A2"]="Licencias Activas"
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:J{fila}")
        excelSheet.append(["Cédula","Nombre","FechaNac","FechaExp","FechaVenc","TipoLicen","TipoSangre","Donador","Sede","Puntaje"])
        for i in range(len(licencia)):
            excelSheet.append([int(licencia[i].getCedula()),licencia[i].getNombre(),licencia[i].getFechaNacimiento(),licencia[i].getFechaExpedicion(),licencia[i].getFechaVencimiento(),licencia[i].getTipoLicencia(),licencia[i].getTipoSangre(),decodificarDonante(licencia[i].getDonador()),licencia[i].getSede(),licencia[i].getPuntaje()])
            if len(licencia[i].getNombre())>len(largoNombre):
                largoNombre=licencia[i].getNombre()
            if len(licencia[i].getSede())>len(largoSede):
                largoSede=licencia[i].getSede()
        for i in range(len(licencia)+1):
            for j in range(1,11):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=excelSheet.column_dimensions["C"].width=excelSheet.column_dimensions["D"].width=excelSheet.column_dimensions["E"].width=14
        excelSheet.column_dimensions["B"].width=len(largoNombre)+4
        excelSheet.column_dimensions["F"].width=len("TipoLicen")+4
        excelSheet.column_dimensions["G"].width=len("TipoSangre")+4
        excelSheet.column_dimensions["H"].width=len("Donador")+4
        excelSheet.column_dimensions["I"].width=len(largoSede)+4
        excelSheet.column_dimensions["J"].width=len("Puntaje")+4
        excel.save("Licencias.xlsx")
        return True
    except:
        return False
#////////////////////Reporte por tipo de licencia
def reporteTipoLicencia(pTipo):
    """
    Funcionamiento:Verificar el tipo de licencia
    Entradas:pTipo(str)Tipo de licencia
    Salidas:NA
    """
    objetos=leer("conductores")
    listaTipoLicencia=[]
    try:
        for objeto in objetos:
            if pTipo in objeto.getTipoLicencia():
                listaTipoLicencia.append(objeto)
        return generarReportePorTipo(pTipo,listaTipoLicencia)
    except:
        return False
def generarReportePorTipo(pTipo,pListaTipoLicencia):
    """
    Funcionamiento:Crear el reporte en excel de los tipos licencias
    Entradas:pTipo(str)tipo de licencia
    pListaTipoLicencia(lista)lista de los tipos de licencia
    Salidas:NA
    """
    try:
        largo="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = "Tipo de licencia"
        excelSheet["A2"]="Tipo de licencia"
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:C{fila}")
        excelSheet.append(["Cédula","Nombre","TipoLicen"])
        for i in range(len(pListaTipoLicencia)):
            excelSheet.append([int(pListaTipoLicencia[i].getCedula()),pListaTipoLicencia[i].getNombre(),pListaTipoLicencia[i].getTipoLicencia()])
            if len(pListaTipoLicencia[i].getNombre())>len(largo):
                largo=pListaTipoLicencia[i].getNombre()
        for i in range(len(pListaTipoLicencia)+1):
            for j in range(1,4):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=14
        excelSheet.column_dimensions["B"].width=len(largo)+4
        excelSheet.column_dimensions["C"].width=len("TipoLicen")+4
        excel.save("TipoLicencia"+pTipo+".xlsx")
        return True
    except:
        return False
#////////////Reporte examen PorSancion//////////////////
def examenPorSancion():
    """
    Funcionamiento:Verificar el puntaje 
    Entradas:NA
    Salidas:NA
    """
    objetos=leer("conductores")
    examen=[]
    try:
        for objeto in objetos:
            if objeto.getPuntaje()>0 and objeto.getPuntaje()<=6:
                examen.append(objeto)
        return examenSancion(examen)
    except:
        return False
def examenSancion(pExamen):
    """
    Funcionamiento:creacion del documento en excel para el examen por sancion
    Entradas:pExamen Las personas que tienen que hacer el examen
    Salidas:NA
    """
    try:
        largo="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = "Examen por Sanción"
        excelSheet["A2"]="Examen por Sanción"
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:D{fila}")
        excelSheet.append(["Cédula","Nombre","TipoLicen","Puntaje"])
        for i in range(len(pExamen)):
            excelSheet.append([int(pExamen[i].getCedula()),pExamen[i].getNombre(),pExamen[i].getTipoLicencia(),pExamen[i].getPuntaje()])
            if len(pExamen[i].getNombre())>len(largo):
                largo=pExamen[i].getNombre()
        for i in range(len(pExamen)+1):
            for j in range(1,5):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=14
        excelSheet.column_dimensions["B"].width=len(largo)+4
        excelSheet.column_dimensions["C"].width=len("TipoLicen")+4
        excelSheet.column_dimensions["D"].width=len("Puntaje")+4
        excel.save("ExamenPorSanción.xlsx")
        return True
    except:
        return False
#------------------------Reporte de los que son donadores de organos-------------------------------------------------
def generarListaDonantesOrganos():
    """
    Funcionamiento:Verificar si el conductor es donador
    Entradas:NA
    Salidas:NA
    """
    objetos=leer("conductores")
    donantes=[]
    try:
        for objeto in objetos:
            if objeto.getDonador():
                donantes.append(objeto)
        return generarReporteDonantesOrganos(donantes)
    except:
        return False
def generarReporteDonantesOrganos(listaDonantes):
    """
    Funcionamiento:Creacion del archivo de excel sobre los donadores de organos
    Entradas:listaDonantes(lista) de los donadores de organos
    Salidas:NA
    """
    try:
        largo="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = "Donantes de organos"
        excelSheet["A2"]="Donantes de órganos"
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:C{fila}")
        excelSheet.append(["Cédula","Nombre","TipoLicen"])
        for i in range(len(listaDonantes)):
            excelSheet.append([int(listaDonantes[i].getCedula()),listaDonantes[i].getNombre(),listaDonantes[i].getTipoLicencia()])#We are assigning values to the row
            if len(listaDonantes[i].getNombre())>len(largo):
                largo=listaDonantes[i].getNombre()
        for i in range(len(listaDonantes)+1):
            for j in range(1,4):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=14
        excelSheet.column_dimensions["B"].width=len(largo)+4
        excelSheet.column_dimensions["C"].width=len("TipoLicen")+4
        excel.save("DonOrg.xlsx")
        return True
    except:
        return False
#------------------------Reporte de las licencias anuladas-------------------------------------------------
def decodificarDonante(booleano):
    """
    Funcionamiento:Decodificacion para saber si una persona es donante
    Entradas:booleano(bool)
    Salidas:NA
    """
    if booleano:
        return "Sí"
    else:
        return "No"
def generarListaAnulados():
    """
    Funcionamiento:Verificar que el puntaje sea igual a 0
    Entradas:NA
    Salidas:NA
    """
    objetos=leer("conductores")
    anulados=[]
    try:
        for objeto in objetos:
            if objeto.getPuntaje()==0:
                anulados.append(objeto)
        return generarReporteAnulados(anulados)
    except:
        return False
def generarReporteAnulados(anulados):
    """
    Funcionamiento:Crear el archivo de excel de los anulados
    Entradas:anulados(licencia con puntaje=0)
    Salidas:NA
    """
    try:
        largoNombre=largoSede="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = "Licencias Anuladas"
        excelSheet["A2"]="Licencias Anuladas"
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:I{fila}")
        excelSheet.append(["Cédula","Nombre","FechaNac","FechaExp","FechaVenc","TipoLicen","TipoSangre","Donador","Sede"])
        for i in range(len(anulados)):
            excelSheet.append([int(anulados[i].getCedula()),anulados[i].getNombre(),anulados[i].getFechaNacimiento(),anulados[i].getFechaExpedicion(),anulados[i].getFechaVencimiento(),anulados[i].getTipoLicencia(),anulados[i].getTipoSangre(),decodificarDonante(anulados[i].getDonador()),anulados[i].getSede()])#We are assigning values to the row
            if len(anulados[i].getNombre())>len(largoNombre):
                largoNombre=anulados[i].getNombre()
            if len(anulados[i].getSede())>len(largoSede):
                largoSede=anulados[i].getSede()
        for i in range(len(anulados)+1):
            for j in range(1,10):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=excelSheet.column_dimensions["C"].width=excelSheet.column_dimensions["D"].width=excelSheet.column_dimensions["E"].width=14
        excelSheet.column_dimensions["B"].width=len(largoNombre)+4
        excelSheet.column_dimensions["F"].width=len("TipoLicen")+4
        excelSheet.column_dimensions["G"].width=len("TipoSangre")+4
        excelSheet.column_dimensions["H"].width=len("Donador")+4
        excelSheet.column_dimensions["I"].width=len(largoSede)+4
        excel.save("Anulados.xlsx")
        return True
    except:
        return False
#------------------------------------Reporte por sede-------------------------------------------
def generarListaPorSede(pSede):
    """
    Funcionamiento:Verificar si la sede se encuentra en el sistema
    Entradas:pSede(str)la sede
    Salidas:NA
    """
    objetos=leer("conductores")
    listaSede=[]
    try:
        for objeto in objetos:
            if objeto.getSede()==pSede:
                listaSede.append(objeto)
        return generarReportePorSede(pSede,listaSede)
    except:
        return False
def generarReportePorSede(pSede,sede):
    """
    Funcionamiento:generar una archivo en excel por el tipo de sede
    Entradas:pSede(str) la sede
    sede
    Salidas:NA
    """
    try:
        largoNombre=largoSede="Nombre"
        excel = Workbook()
        excelSheet= excel.active 
        excelSheet.title = pSede
        excelSheet["A2"]="Licencias de la sede: "+pSede
        excelSheet["A3"]=str(datetime.datetime.now())[:-7]
        for fila in range(2,4):
            excelSheet[f"A{fila}"].font=Font(bold=True,size=10,name="Times New Roman")
            excelSheet[f"A{fila}"].alignment=Alignment(horizontal="center")
            excelSheet[f"A{fila}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
            excelSheet.merge_cells(f"A{fila}:J{fila}")
        excelSheet.append(["Cédula","Nombre","FechaNac","FechaExp","FechaVenc","TipoLicen","TipoSangre","Donador","Sede","Puntaje"])
        for i in range(len(sede)):
            excelSheet.append([int(sede[i].getCedula()),sede[i].getNombre(),sede[i].getFechaNacimiento(),sede[i].getFechaExpedicion(),sede[i].getFechaVencimiento(),sede[i].getTipoLicencia(),sede[i].getTipoSangre(),decodificarDonante(sede[i].getDonador()),sede[i].getSede(),sede[i].getPuntaje()])
            if len(sede[i].getNombre())>len(largoNombre):
                largoNombre=sede[i].getNombre()
            if len(sede[i].getSede())>len(largoSede):
                largoSede=sede[i].getSede()
        for i in range(len(sede)+1):
            for j in range(1,11):
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].font=Font(size=10,name="Time New Roman")
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].border=Border(left=Side(border_style="medium",color='000000'),right=Side(border_style="medium",color='000000'),top=Side(border_style="medium",color='000000'),bottom=Side(border_style="medium",color='000000'))
                excelSheet[f"{str(get_column_letter(j))}{str(i+4)}"].alignment=Alignment(horizontal="center")
        excelSheet.column_dimensions["A"].width=excelSheet.column_dimensions["C"].width=excelSheet.column_dimensions["D"].width=excelSheet.column_dimensions["E"].width=14
        excelSheet.column_dimensions["B"].width=len(largoNombre)+4
        excelSheet.column_dimensions["F"].width=len("TipoLicen")+4
        excelSheet.column_dimensions["G"].width=len("TipoSangre")+4
        excelSheet.column_dimensions["H"].width=len("Donador")+4
        excelSheet.column_dimensions["I"].width=len(largoSede)+4
        excelSheet.column_dimensions["J"].width=len("Puntaje")+4
        excel.save(pSede+".xlsx")
        return True
    except:
        return False